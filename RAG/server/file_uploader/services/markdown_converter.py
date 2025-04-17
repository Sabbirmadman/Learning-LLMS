import os
from markitdown import MarkItDown
import mammoth
import openpyxl
import pytesseract
from PIL import Image
import csv
from charset_normalizer import from_path
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Adjust path if needed

# extracters
def extract_markdown_from_pdf(file_path):
    try:
        md = MarkItDown()
        result = md.convert(file_path)
        if result.text_content.strip():
            return result.text_content
    except Exception as e:
        print(f"Failed to convert PDF to Markdown for {file_path}: {e}")
    return None

def extract_markdown_from_docx(file_path):
    try:
        with open(file_path, "rb") as docx_file:
            result = mammoth.convert_to_markdown(docx_file)
            if result.value.strip():
                return result.value
    except Exception as e:
        print(f"Failed to convert DOCX to Markdown for {file_path}: {e}")
    return None

def extract_markdown_from_xlsx(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        markdown = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            markdown += f"## {sheet}\n\n"
            for row in worksheet.iter_rows(values_only=True):
                row_md = "| " + " | ".join([str(cell) if cell is not None else "" for cell in row]) + " |\n"
                markdown += row_md
            markdown += "\n"
        if markdown.strip():
            return markdown
    except Exception as e:
        print(f"Failed to convert XLSX to Markdown for {file_path}: {e}")
    return None

def extract_markdown_from_image(file_path):
    try:
        image = Image.open(file_path)
        text = pytesseract.image_to_string(image)
        if text.strip():
            return f"![Image]({file_path})\n\n" + text
    except Exception as e:
        print(f"Failed to extract text from image {file_path}: {e}")
    return None

def extract_markdown_from_csv(file_path):
    try:
        # Detect file encoding using charset-normalizer
        result = from_path(file_path).best()
        if not result:
            print(f"Failed to detect encoding for {file_path}")
            return None
        
        encoding = result.encoding
        print(f"Detected encoding for {file_path}: {encoding}")
        
        # Read the CSV file with the detected encoding - limited to first 20 lines
        with open(file_path, 'r', newline='', encoding=encoding) as csvfile:
            csvreader = csv.reader(csvfile)
            # Get only first 20 rows
            rows = []
            for i, row in enumerate(csvreader):
                rows.append(row)
                if i >= 50:  # Stop after reading 20 rows (0-19)
                    break
                    
            if not rows:
                return None
            
            markdown = []
            # Add header row
            headers = rows[0]
            markdown.append("| " + " | ".join(headers) + " |")
            
            # Add separator row
            markdown.append("| " + " | ".join(["---"] * len(headers)) + " |")
            
            # Add data rows (limited to what we've read)
            for row in rows[1:]:
                markdown.append("| " + " | ".join(str(cell) for cell in row) + " |")
            
            markdown_str = "\n".join(markdown)
            return markdown_str if markdown_str.strip() else None
    except Exception as e:
        print(f"Failed to convert CSV to Markdown {file_path}: {e}")
    return None